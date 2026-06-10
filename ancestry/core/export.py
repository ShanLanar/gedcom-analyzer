"""
Export-Funktionen: CSV und XLSX (Matches + Shared Matches).
"""

import csv
import logging
from typing import Optional

from ancestry.models import DnaMatch, SharedMatch

log = logging.getLogger(__name__)

MATCH_COLUMNS = [
    "display_name", "gender", "shared_cm", "shared_segments", "longest_segment",
    "predicted_relationship", "confidence", "relationship_range",
    "has_tree", "tree_status", "tree_size", "has_common_ancestor",
    "starred", "note", "custom_relationship", "paternal_maternal",
    "ethnicity_regions", "last_login", "fetched_at", "match_guid",
]

SHARED_COLUMNS = [
    "match_guid_a", "display_name_b", "shared_cm_b", "shared_cm_ab",
    "shared_segments_b", "relationship_b", "has_tree_b",
    "match_guid_b", "fetched_at",
]

MATCH_LABELS = {
    "display_name"           : "Name",
    "gender"                 : "Geschlecht",
    "shared_cm"              : "Gemeinsame cM",
    "shared_segments"        : "Segmente",
    "longest_segment"        : "Längstes Segment (cM)",
    "predicted_relationship" : "Beziehung (Ancestry)",
    "confidence"             : "Konfidenz",
    "relationship_range"     : "Beziehungsbereich",
    "has_tree"               : "Hat Stammbaum",
    "tree_status"            : "Stammbaum-Status",
    "tree_size"              : "Stammbaum-Personen",
    "has_common_ancestor"    : "Gemeinsamer Vorfahre",
    "starred"                : "Markiert",
    "note"                   : "Notiz",
    "custom_relationship"    : "Eigene Beziehung",
    "paternal_maternal"      : "Seite (v/m)",
    "ethnicity_regions"      : "Herkunftsregionen",
    "last_login"             : "Letzter Login",
    "fetched_at"             : "Abgerufen am",
    "match_guid"             : "Match-GUID",
}

SHARED_LABELS = {
    "match_guid_a"    : "Primärer Match (GUID)",
    "display_name_b"  : "Shared Match (Name)",
    "shared_cm_b"     : "cM (Benutzer ↔ Shared)",
    "shared_cm_ab"    : "cM (Primär ↔ Shared)",
    "shared_segments_b": "Segmente (Shared)",
    "relationship_b"  : "Beziehung (Shared)",
    "has_tree_b"      : "Stammbaum (Shared)",
    "match_guid_b"    : "Shared Match (GUID)",
    "fetched_at"      : "Abgerufen am",
}


def _fmt(val) -> str:
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    if isinstance(val, bool):
        return "Ja" if val else "Nein"
    return str(val) if val is not None else ""


# ── CSV ───────────────────────────────────────────────────────────────────────

def export_csv(matches: list[DnaMatch], filepath: str) -> int:
    headers = [MATCH_LABELS.get(c, c) for c in MATCH_COLUMNS]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=headers, quoting=csv.QUOTE_ALL)
        w.writeheader()
        for m in matches:
            d = m.to_dict()
            w.writerow({MATCH_LABELS.get(c, c): _fmt(d.get(c, ""))
                        for c in MATCH_COLUMNS})
    log.info("CSV-Export Matches: %d → %s", len(matches), filepath)
    return len(matches)


def export_shared_csv(shared: list[SharedMatch], filepath: str,
                       match_name_map: Optional[dict] = None) -> int:
    """
    Exportiert Shared Matches als CSV.
    match_name_map: {match_guid_a: display_name} – optional für lesbaren Namen.
    """
    headers = ["Primärer Match (Name)"] + [SHARED_LABELS.get(c, c) for c in SHARED_COLUMNS]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=headers, quoting=csv.QUOTE_ALL)
        w.writeheader()
        for sm in shared:
            d = sm.to_dict()
            row = {"Primärer Match (Name)":
                   (match_name_map or {}).get(sm.match_guid_a, "")}
            for c in SHARED_COLUMNS:
                row[SHARED_LABELS.get(c, c)] = _fmt(d.get(c, ""))
            w.writerow(row)
    log.info("CSV-Export Shared Matches: %d → %s", len(shared), filepath)
    return len(shared)


# ── XLSX ──────────────────────────────────────────────────────────────────────

def _write_xlsx_sheet(ws, headers: list[str], rows: list[list], col_widths: list[int]):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl nicht installiert: pip install openpyxl")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill    = PatternFill("solid", fgColor="EBF0FA")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    for row_idx, row_data in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_xlsx(matches: list[DnaMatch], filepath: str,
                shared: Optional[list[SharedMatch]] = None,
                match_name_map: Optional[dict] = None,
                gedcom_summary: Optional[list[dict]] = None,
                stats: Optional[dict] = None,
                analysis: Optional[list[dict]] = None) -> int:
    """
    Exportiert Matches (und optional Shared Matches + GEDCOM-Vergleich + Statistik
    + Herkunft/Seiten) in eine XLSX-Datei.
    gedcom_summary: Ausgabe von bridge.get_gedcom_relationship_summary() – optional.
    stats:          Ausgabe von database.get_statistics() – optional (eigenes Blatt).
    analysis:       Liste {name,cm,side,origin_rule,origin_ml} – optional (Blatt).
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl nicht installiert: pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── Blatt 0: Statistik (optional, zuerst) ─────────────────────────────────
    if stats:
        ws0 = wb.active
        ws0.title = "Statistik"
        kz = [
            ("Matches gesamt",            stats.get("total")),
            ("Höchste gem. cM",           stats.get("max_cm")),
            ("Durchschn. cM",             round(stats["avg_cm"], 1) if stats.get("avg_cm") else None),
            ("Markiert (Stern)",          stats.get("starred_count")),
            ("Mit Stammbaum",             stats.get("with_tree")),
            ("Mit Notiz",                 stats.get("with_note")),
            ("Shared-Match-Einträge",     stats.get("shared_total")),
            ("Matches mit Shared-Daten",  stats.get("shared_primary_count")),
            ("Ahnentafeln geladen",       stats.get("ped_loaded")),
            ("Verschiedene Pedigree-Nachnamen", stats.get("ped_surnames")),
            ("Ø Ahnentafel-Tiefe (Generationen)", stats.get("ped_avg_depth")),
            ("GEDCOM-Personen",           stats.get("gedcom_persons")),
            ("GEDCOM-verknüpfte Matches", stats.get("gedcom_linked")),
        ]
        rows0 = [[k, _fmt(v if v is not None else "")] for k, v in kz]
        _write_xlsx_sheet(ws0, ["Kennzahl", "Wert"], rows0, [34, 16])
        # Beziehungs-Aufschlüsselung
        rb = stats.get("relationship_breakdown") or []
        if rb:
            start = len(rows0) + 3
            ws0.cell(row=start, column=1, value="Beziehung").font = \
                __import__("openpyxl").styles.Font(bold=True)
            ws0.cell(row=start, column=2, value="Anzahl").font = \
                __import__("openpyxl").styles.Font(bold=True)
            for i, (rel, cnt) in enumerate(rb, start=start + 1):
                ws0.cell(row=i, column=1, value=rel)
                ws0.cell(row=i, column=2, value=cnt)
        # Kit-Aufschlüsselung
        kb = stats.get("kit_breakdown") or []
        if kb:
            col = 4
            ws0.cell(row=1, column=col, value="Kit").font = \
                __import__("openpyxl").styles.Font(bold=True)
            ws0.cell(row=1, column=col + 1, value="Matches").font = \
                __import__("openpyxl").styles.Font(bold=True)
            for i, (kit, cnt) in enumerate(kb, start=2):
                ws0.cell(row=i, column=col, value=kit)
                ws0.cell(row=i, column=col + 1, value=cnt)
        ws1 = wb.create_sheet("DNA-Matches")
    else:
        ws1 = wb.active
        ws1.title = "DNA-Matches"

    # ── Blatt 1: Matches ──────────────────────────────────────────────────────
    headers1 = [MATCH_LABELS.get(c, c) for c in MATCH_COLUMNS]
    widths1  = [30, 14, 10, 18, 25, 12, 20, 12, 12, 10, 40, 20, 30, 20, 12, 20, 20, 36]
    rows1 = []
    for m in matches:
        d = m.to_dict()
        rows1.append([_fmt(d.get(c, "")) for c in MATCH_COLUMNS])
    _write_xlsx_sheet(ws1, headers1, rows1, widths1)

    # ── Blatt 2: Shared Matches (optional) ────────────────────────────────────
    if shared:
        ws2 = wb.create_sheet("Shared Matches")
        headers2 = ["Primärer Match"] + [SHARED_LABELS.get(c, c) for c in SHARED_COLUMNS]
        widths2  = [30, 36, 20, 18, 14, 16, 22, 12, 36, 20]
        rows2 = []
        for sm in shared:
            d = sm.to_dict()
            name_a = (match_name_map or {}).get(sm.match_guid_a, "")
            rows2.append([name_a] + [_fmt(d.get(c, "")) for c in SHARED_COLUMNS])
        _write_xlsx_sheet(ws2, headers2, rows2, widths2)

    # ── Blatt 3: GEDCOM-Verwandtschafts-Vergleich (optional) ─────────────────
    if gedcom_summary:
        ws3 = wb.create_sheet("GEDCOM-Vergleich")
        headers3 = [
            "Name", "Gemeinsame cM",
            "Ancestry-Beziehung", "GEDCOM-Beziehung", "Multiplikator",
            "Anzahl Verknüpfungen", "Treffer-Score",
            "Gemeinsamer Ahne (GEDCOM)", "Geburtsjahr Ahne",
            "Tiefe im eigenen Baum", "Tiefe beim Match",
        ]
        widths3 = [30, 14, 22, 22, 12, 20, 14, 35, 14, 20, 20]
        rows3 = [
            [
                r.get("display_name", ""),
                r.get("shared_cm", ""),
                r.get("ancestry_rel", ""),
                r.get("ged_relationship", ""),
                r.get("multiplier", ""),
                r.get("link_count", ""),
                r.get("best_score", ""),
                r.get("ged_common_ancestor", ""),
                r.get("ged_ancestor_year", ""),
                r.get("root_gen_depth", ""),
                r.get("match_gen_depth", ""),
            ]
            for r in gedcom_summary
        ]
        _write_xlsx_sheet(ws3, headers3, rows3, widths3)

    # ── Blatt 4: Herkunft & Seiten (optional) ────────────────────────────────
    if analysis:
        ws4 = wb.create_sheet("Herkunft & Seiten")
        headers4 = ["Name", "Gem. cM", "Seite",
                    "Herkunft (Regel)", "Herkunft (ML)"]
        widths4 = [30, 12, 14, 28, 28]
        rows4 = [[a.get("name", ""), _fmt(a.get("cm", "")), a.get("side", ""),
                  a.get("origin_rule", ""), a.get("origin_ml", "")]
                 for a in analysis]
        _write_xlsx_sheet(ws4, headers4, rows4, widths4)

    wb.save(filepath)
    log.info("XLSX-Export: %d Matches, %d Shared, %d GEDCOM-Links, "
             "Statistik=%s, Analyse=%d → %s",
             len(matches), len(shared) if shared else 0,
             len(gedcom_summary) if gedcom_summary else 0,
             "ja" if stats else "nein", len(analysis) if analysis else 0, filepath)
    return len(matches)
