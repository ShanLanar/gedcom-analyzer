# -*- coding: utf-8 -*-
"""tasks/export.py – Excel- und JSON-Export aller Analyse-Sheets"""

import html
import json
import os
from datetime import datetime


# ── Excel-Export ───────────────────────────────────────────────────────────────

def export_to_excel(all_sheets: list, output_path: str,
                    progress_cb=None) -> bool:
    """
    Schreibt eine Excel-Datei mit einem Sheet pro Eintrag in all_sheets.
    all_sheets = [(sheet_name, [headers], [rows]), ...]

    Nutzt openpyxl im write_only-Modus: Cells werden gestreamt geschrieben,
    Memory bleibt flach, und der Export ist bei großen Sheets etwa 10×
    schneller als der Default-Modus mit Banding.
    """
    p = progress_cb or (lambda m, **kw: None)
    p(f"Erstelle Excel: {output_path} …")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.cell.cell import WriteOnlyCell
        from openpyxl.utils import get_column_letter
    except ImportError:
        p("openpyxl nicht verfügbar – bitte installieren: pip install openpyxl",
          tag="err")
        return False

    wb = Workbook(write_only=True)

    header_fill  = PatternFill(start_color="366092", end_color="366092",
                                fill_type="solid")
    header_font  = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    used_titles: set = set()

    def _unique_title(name: str) -> str:
        base = name[:31]
        if base not in used_titles:
            used_titles.add(base)
            return base
        for n in range(2, 100):
            suffix = f" ({n})"
            cand = base[: 31 - len(suffix)] + suffix
            if cand not in used_titles:
                used_titles.add(cand)
                return cand
        used_titles.add(base)
        return base

    CAP = 200_000
    for sheet_name, headers, data in all_sheets:
        if not data:
            continue
        try:
            ws = wb.create_sheet(title=_unique_title(sheet_name))

            # Spaltenbreiten anhand der Headers + ersten 50 Datenzeilen
            # bestimmen — muss vor dem ersten append() passieren.
            sample = data[:50]
            for ci, header in enumerate(headers, start=1):
                max_len = len(str(header))
                for row in sample:
                    if ci - 1 < len(row) and row[ci - 1] is not None:
                        v = str(row[ci - 1])
                        if len(v) > max_len:
                            max_len = len(v)
                ws.column_dimensions[get_column_letter(ci)].width = min(
                    max_len + 2, 50)
            ws.freeze_panes = "A2"

            # Header mit Styling
            styled = []
            for h in headers:
                c = WriteOnlyCell(ws, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = header_align
                styled.append(c)
            ws.append(styled)

            # Datenzeilen streamen
            n_rows = 0
            for ri, row in enumerate(data[:CAP], start=2):
                ws.append(row)
                n_rows += 1
                if n_rows % 10_000 == 0:
                    p(f"    {n_rows:,}/{min(len(data), CAP):,} Zeilen …")
            p(f"  ✓ Sheet '{ws.title}': {n_rows:,} Zeilen")

        except Exception as e:
            p(f"Fehler bei Sheet '{sheet_name}': {e}", tag="warn")

    p("Speichere Excel-Datei (kann bei großen Workbooks etwas dauern) …")
    tmp_path = output_path + ".tmp"
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, output_path)   # atomarer Rename – kein halbfertiges File
        size_mb = os.path.getsize(output_path) / 1_048_576
        p(f"Excel gespeichert: {output_path} ({size_mb:.1f} MB)", tag="ok")
        return True
    except Exception as e:
        if os.path.exists(tmp_path):
            try: os.unlink(tmp_path)
            except OSError: pass
        p(f"Fehler beim Speichern von Excel: {e}", tag="err")
        return False


# ── JSON-Export ────────────────────────────────────────────────────────────────

def export_to_json(data: dict, output_path: str, progress_cb=None) -> bool:
    p = progress_cb or (lambda m, **kw: None)
    p(f"JSON-Export: {output_path} …")
    try:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        size_mb = os.path.getsize(output_path) / 1_048_576
        p(f"JSON gespeichert: {output_path} ({size_mb:.1f} MB)", tag="ok")
        return True
    except Exception as e:
        p(f"Fehler beim JSON-Export: {e}", tag="err")
        return False


# ── Sheet-Definitionen ─────────────────────────────────────────────────────────

def build_symbol_sheet(individuals) -> tuple:
    return ("Symbol-Statistik", ["Symbol", "Bedeutung", "Anzahl"], [
        ["✠", "Deutsche Soldaten",
         sum(1 for i in individuals.values() if i.get("GERMAN_SOLDIER"))],
        ["★", "Andere Soldaten",
         sum(1 for i in individuals.values() if i.get("OTHER_SOLDIER"))],
        ["⚔", "Gefallene",
         sum(1 for i in individuals.values() if i.get("DIED_IN_BATTLE"))],
        ["‡", "Linie endet",
         sum(1 for i in individuals.values() if i.get("LINE_ENDS"))],
        ["mig.", "Migriert (markiert)",
         sum(1 for i in individuals.values() if i.get("MIGRATED"))],
    ])


def build_gedcom_events_sheet(individuals) -> tuple:
    return ("GEDCOM-Events", ["Event-Typ", "Anzahl", "Beschreibung"], [
        ["EMIG",
         sum(1 for i in individuals.values() if i.get("EMIG", {}).get("DATE")),
         "Auswanderung (EMIG Event)"],
        ["IMMI",
         sum(1 for i in individuals.values() if i.get("IMMI", {}).get("DATE")),
         "Einwanderung (IMMI Event)"],
        ["mig. im Namen",
         sum(1 for i in individuals.values()
             if "mig." in (i.get("NAME") or "").lower()),
         "mig. im Namen markiert"],
    ])


# ── HTML-Übersicht ─────────────────────────────────────────────────────────────

def export_html_overview(state: dict, output_path: str,
                          progress_cb=None) -> bool:
    """Schreibt eine selbsterklärende HTML-Datei mit den wichtigsten
    Kennzahlen. Keine externen Abhängigkeiten."""
    p = progress_cb or (lambda m, **kw: None)
    p(f"HTML-Übersicht: {output_path} …")

    indi  = state.get("individuals", {}) or {}
    fams  = state.get("families", {}) or {}
    stats = state.get("comprehensive_stats", []) or []
    surn  = (state.get("surname_results", []) or [])[:20]
    coun  = (state.get("country_dist_results", []) or [])[:10]
    waves = state.get("migration_waves", []) or []
    dem   = state.get("demographic_results", []) or []

    def esc(v) -> str:
        return html.escape(str(v))

    def table(headers, rows) -> str:
        head = "".join(f"<th>{esc(h)}</th>" for h in headers)
        body = "".join(
            "<tr>" + "".join(f"<td>{esc(c)}</td>" for c in r) + "</tr>"
            for r in rows)
        return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"

    sections = []
    if stats:
        sections.append(("Gesamtstatistik",
                          table(["Kennzahl", "Wert", "Anteil"], stats)))
    if dem:
        from tasks.demographics import DEMOGRAPHIC_HEADERS
        sections.append(("Demografie pro Epoche",
                          table(DEMOGRAPHIC_HEADERS, dem)))
    if surn:
        from tasks.demographics import SURNAME_HEADERS
        sections.append(("Top-20 Familiennamen",
                          table(SURNAME_HEADERS, surn)))
    if coun:
        from tasks.demographics import COUNTRY_HEADERS
        sections.append(("Top-10 Geburtsländer",
                          table(COUNTRY_HEADERS, coun)))
    if waves:
        from tasks.migration import WAVES_HEADERS
        sections.append(("Migrationswellen", table(WAVES_HEADERS, waves)))

    body_html = "\n".join(
        f"<section><h2>{esc(title)}</h2>{tbl}</section>"
        for title, tbl in sections)

    doc = f"""<!DOCTYPE html>
<html lang="de">
<head>
  <meta charset="utf-8">
  <title>Ahnen-Analyse Übersicht</title>
  <style>
    :root {{
      --bg: #1e1e2e; --bg2: #2a2a3e; --accent: #7c7cf8; --fg: #cdd6f4;
      --fg-dim: #6c7086; --row-alt: #232336;
    }}
    body {{ font-family: 'Segoe UI', sans-serif; max-width: 1100px;
            margin: 2em auto; padding: 0 1em; background: var(--bg);
            color: var(--fg); }}
    h1 {{ color: var(--accent); border-bottom: 2px solid var(--accent);
          padding-bottom: 0.3em; }}
    h2 {{ color: var(--accent); margin-top: 2em; }}
    .meta {{ color: var(--fg-dim); font-size: 0.9em; }}
    table {{ border-collapse: collapse; width: 100%; margin-top: 0.5em;
              font-size: 0.9em; }}
    th, td {{ border: 1px solid var(--bg2); padding: 4px 8px;
              text-align: left; }}
    th {{ background: var(--accent); color: #fff; }}
    tr:nth-child(even) {{ background: var(--row-alt); }}
    section {{ margin-bottom: 2em; }}
  </style>
</head>
<body>
  <h1>🧬 Ahnen-Analyse Übersicht</h1>
  <p class="meta">Stand: {esc(datetime.now().strftime('%Y-%m-%d %H:%M'))}
      &middot; {len(indi):,} Personen &middot; {len(fams):,} Familien</p>
  {body_html or '<p class="meta">Keine Analyseergebnisse vorhanden.</p>'}
</body>
</html>"""

    try:
        d = os.path.dirname(output_path)
        if d and not os.path.exists(d):
            os.makedirs(d, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(doc)
        size_kb = os.path.getsize(output_path) / 1024
        p(f"HTML gespeichert: {output_path} ({size_kb:.1f} KB)", tag="ok")
        return True
    except OSError as e:
        p(f"Fehler beim HTML-Export: {e}", tag="err")
        return False


def export_timeline_html(individuals: dict, families: dict, output_path: str,
                          root_related_ids: set | None = None,
                          progress_cb=None) -> bool:
    """Chronologische HTML-Zeitlinie aller Ereignisse aller (oder verwandter) Personen."""
    p = progress_cb or (lambda m, **kw: None)
    p(f"Timeline-HTML: {output_path} …")

    scope = root_related_ids if root_related_ids else set(individuals)
    events: list = []
    for iid in scope:
        indi = individuals.get(iid)
        if not indi:
            continue
        name = html.escape((indi.get("NAME") or "Unbekannt")[:50])
        sex  = indi.get("SEX", "U")
        for ev_tag, ev_label in (("BIRT", "Geburt"), ("DEAT", "Tod"),
                                  ("EMIG", "Auswanderung"), ("IMMI", "Einwanderung")):
            ev = indi.get(ev_tag) or {}
            year = ev.get("YEAR")
            if not year:
                continue
            place = html.escape((ev.get("PLAC") or "")[:40])
            events.append((year, ev_label, iid, name, sex, place))

    events.sort(key=lambda e: e[0])
    p(f"  {len(events):,} Ereignisse sortiert …")

    _sex_color = {"M": "#7eb8f7", "F": "#f7a8c4", "U": "#b0b0b0"}
    rows_html = "\n".join(
        f'<tr><td>{year}</td><td class="ev-{ev_label[:3].lower()}">{ev_label}</td>'
        f'<td style="color:{_sex_color.get(sex, "#b0b0b0")}">{name}</td>'
        f'<td><small>{place}</small></td></tr>'
        for year, ev_label, iid, name, sex, place in events
    )
    doc = f"""<!DOCTYPE html>
<html lang="de">
<head>
  <meta charset="utf-8">
  <title>Ahnen-Zeitlinie</title>
  <style>
    :root {{--bg:#1e1e2e;--bg2:#2a2a3e;--accent:#7c7cf8;--fg:#cdd6f4;--dim:#6c7086;}}
    body {{font-family:'Segoe UI',sans-serif;max-width:1100px;margin:2em auto;
           padding:0 1em;background:var(--bg);color:var(--fg);}}
    h1 {{color:var(--accent);border-bottom:2px solid var(--accent);padding-bottom:.3em;}}
    input {{background:var(--bg2);color:var(--fg);border:1px solid var(--accent);
            padding:4px 8px;width:220px;border-radius:4px;font-size:.9em;}}
    table {{border-collapse:collapse;width:100%;margin-top:.8em;font-size:.88em;}}
    th,td {{border:1px solid var(--bg2);padding:4px 8px;text-align:left;}}
    th {{background:var(--accent);color:#fff;position:sticky;top:0;z-index:1;}}
    tr:nth-child(even) {{background:#232336;}}
    .ev-geb {{color:#50fa7b;}} .ev-tod {{color:#ff5555;}}
    .ev-aus {{color:#ffb86c;}} .ev-imm {{color:#7c7cf8;}}
    .meta {{color:var(--dim);font-size:.9em;}}
    #filter {{margin:.6em 0;}}
  </style>
</head>
<body>
  <h1>🕰 Ahnen-Zeitlinie</h1>
  <p class="meta">{len(events):,} Ereignisse · {len(scope):,} Personen ·
    Stand: {html.escape(datetime.now().strftime('%Y-%m-%d %H:%M'))}</p>
  <div id="filter">
    <input type="text" id="q" placeholder="Person oder Ort filtern …" oninput="filterTable()">
  </div>
  <table id="tbl">
    <thead><tr><th>Jahr</th><th>Ereignis</th><th>Person</th><th>Ort</th></tr></thead>
    <tbody id="tbody">{rows_html}</tbody>
  </table>
  <script>
    function filterTable(){{
      const q=document.getElementById('q').value.toLowerCase();
      document.querySelectorAll('#tbody tr').forEach(r=>{{
        r.style.display=r.textContent.toLowerCase().includes(q)?'':'none';
      }});
    }}
  </script>
</body>
</html>"""
    try:
        d = os.path.dirname(output_path)
        if d and not os.path.exists(d):
            os.makedirs(d, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(doc)
        size_kb = os.path.getsize(output_path) / 1024
        p(f"Timeline gespeichert: {output_path} ({size_kb:.1f} KB)", tag="ok")
        return True
    except OSError as e:
        p(f"Fehler beim Timeline-Export: {e}", tag="err")
        return False


def build_location_info_sheet(location_data: dict) -> tuple:
    countries = len(location_data.get("countries", {}))
    states = sum(len(c.get("states", {}))
                 for c in location_data.get("countries", {}).values())
    return ("Ortsdaten-Info", ["Kategorie", "Anzahl", "Beschreibung"], [
        ["Länder", countries, "Anzahl der definierten Länder"],
        ["Bundesstaaten", states, "Gesamtzahl der Bundesstaaten/Provinzen"],
        ["Bezirks-Indikatoren",
         len(location_data.get("district_indicators", [])),
         "Wörter zur Bezirkserkennung"],
        ["Provinz-Indikatoren",
         len(location_data.get("province_indicators", [])),
         "Wörter zur Provinz/Region-Erkennung"],
    ])
